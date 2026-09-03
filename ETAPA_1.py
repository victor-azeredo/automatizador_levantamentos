import os
import glob
import json
import re
import pandas as pd
import pdfplumber
from thefuzz import process, fuzz
from google import genai
from google.genai import types
from openpyxl import load_workbook
from dotenv import load_dotenv

# ==========================================
# 0. CONFIGURAÇÕES E PASTAS
# ==========================================
LIMITE_SIMILARIDADE = 80 # Define o quão parecido o nome tem que ser (0 a 100)

pasta_checklist = r"C:\Users\joao\OneDrive\Desktop\AUTOMATIZADOR_LEVANTAMENTOS\COLOQUE_AQUI_CHECKLIST"
pasta_base = r"C:\Users\joao\OneDrive\Desktop\AUTOMATIZADOR_LEVANTAMENTOS\PLANILHA_BASE"
pasta_pos = r"C:\Users\joao\OneDrive\Desktop\AUTOMATIZADOR_LEVANTAMENTOS\POs"

try:
    arquivo_checklist = glob.glob(os.path.join(pasta_checklist, "*.xlsx"))[0]
    arquivo_destino = glob.glob(os.path.join(pasta_base, "*.xlsx"))[0]
    arquivos_pdf_pos = glob.glob(os.path.join(pasta_pos, "*.pdf"))
except IndexError:
    print("Erro: Arquivos base ou checklist não encontrados nas pastas!")
    exit()

# ==========================================
# 1. EXTRAÇÃO DO CHECKLIST
# ==========================================
print("1. Lendo checklist...")
df_checklist = pd.read_excel(arquivo_checklist, sheet_name='Sheet1', skiprows=6)
df_checklist.columns = ['Item', 'Qtd_Necessaria', 'Unidade', 'Qtd_Conferida', 'OK']

# Pega apenas itens com quantidade > 0
df_filtrado = df_checklist[pd.to_numeric(df_checklist['Qtd_Necessaria'], errors='coerce') > 0].copy()

lista_para_ia = []
for index, row in df_filtrado.iterrows():
    lista_para_ia.append({
        "nome": str(row['Item']),
        "quantidade": int(row['Qtd_Necessaria']),
        "tipo": "" 
    })

payload_json = json.dumps(lista_para_ia, ensure_ascii=False, indent=2)

# ==========================================
# 2. CLASSIFICAÇÃO COM A IA (Gemini)
# ==========================================
print("2. Classificando Equipamentos e Materiais com a IA...")
load_dotenv()
api_key = os.getenv("API_KEY")

client = genai.Client(api_key=api_key)

prompt = f"""
Você é um classificador de materiais. Analise a lista JSON abaixo. 
Para cada item, preencha a chave "tipo" com:
"E" se for Equipamento.
"M" se for Material de consumo/instalação.
Retorne APENAS o JSON preenchido.

Lista:
{payload_json}
"""

config = types.GenerateContentConfig(response_mime_type="application/json")
resposta = client.models.generate_content(model='gemini-3.5-flash', contents=prompt, config=config)
json_classificado = json.loads(resposta.text.strip())

# Separando as listas
lista_m = [i for i in json_classificado if i['tipo'] == 'M']
lista_e = [i for i in json_classificado if i['tipo'] == 'E']

# ==========================================
# 3. EXTRAÇÃO DAS POs (Leitura dos PDFs)
# ==========================================
print("3. Lendo arquivos de PO e extraindo valores...")
dicionario_pos = {}

for pdf_path in arquivos_pdf_pos:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = "".join([pagina.extract_text() + "\n" for pagina in pdf.pages])
            
            for linha in texto_completo.split('\n'):
                # Regex para capturar Padrão OLS: Nome + Preço Unitário + Qtd + Preço Total
                match = re.search(r'(.*?)\s+([\d\s]+,\d{2})\s+\d+\s+[\d\s]+,\d{2}$', linha)
                if match:
                    nome_item_po = match.group(1).strip()
                    valor_float = float(match.group(2).replace(' ', '').replace(',', '.'))
                    dicionario_pos[nome_item_po] = valor_float
    except Exception as e:
        print(f"Erro ao ler {os.path.basename(pdf_path)}: {e}")

# ==========================================
# 4. CRUZAMENTO DE DADOS (Fuzzy Matching)
# ==========================================
print("4. Cruzando dados do checklist com as POs...")
nomes_pos_disponiveis = list(dicionario_pos.keys())

def buscar_preco(nome_item):
    if not nomes_pos_disponiveis: return None
    melhor_match = process.extractOne(nome_item, nomes_pos_disponiveis, scorer=fuzz.token_sort_ratio)
    if melhor_match and melhor_match[1] >= LIMITE_SIMILARIDADE:
        return dicionario_pos[melhor_match[0]]
    return None # Nulo se for item da casa/estoque

# Adicionando os preços às listas
for item in lista_m: item['preco_unitario'] = buscar_preco(item['nome'])
for item in lista_e: item['preco_unitario'] = buscar_preco(item['nome'])

# ==========================================
# 5. ESCRITA NA PLANILHA (O ÚNICO LOAD)
# ==========================================
print("5. Escrevendo os resultados na planilha genérica...")
wb = load_workbook(arquivo_destino)
aba_materiais = wb['Materiais']
aba_equipamentos = wb['Equipamentos']

linha_inicio = 3 # Pula os cabeçalhos

# Escrevendo Materiais
for i, mat in enumerate(lista_m):
    aba_materiais.cell(row=linha_inicio + i, column=2, value=mat['nome'])       
    aba_materiais.cell(row=linha_inicio + i, column=3, value=mat['quantidade'])
    # Se na aba de materiais a coluna de valor for diferente da 4, ajuste o número abaixo:
    aba_materiais.cell(row=linha_inicio + i, column=4, value=mat['preco_unitario']) 

# Escrevendo Equipamentos
for i, eqp in enumerate(lista_e):
    aba_equipamentos.cell(row=linha_inicio + i, column=2, value=eqp['nome'])       
    aba_equipamentos.cell(row=linha_inicio + i, column=3, value=eqp['quantidade']) 
    aba_equipamentos.cell(row=linha_inicio + i, column=4, value=eqp['preco_unitario']) # Custo Un.

caminho_final = r"C:\Users\joao\OneDrive\Desktop\AUTOMATIZADOR_LEVANTAMENTOS\PLANILHA CRUZADA\LEVANTAMENTO_PREENCHIDO.xlsx"
wb.save(caminho_final)

print(f"Sucesso Total! Planilha final salva em:\n{caminho_final}")